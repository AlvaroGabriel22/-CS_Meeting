"""Raw-data fixtures for the parser tests.

⚠ **These are provisional synthetic fixtures.**  The real IQC / OQC / FIELD
workbooks were not available when Sprint 0 was written, so these files
reproduce the structures *described* in the specification — three-level
hierarchies (``SEC`` → ``Total``/``TSI``/``Packing`` → ``PPM``/``Def.``/``Insp.``
and ``ASR`` → ``MX``/``Mobile`` → ``Target``/``Result``), merged cells, mixed
year/month/week headers, ``NA``, ``#DIV/0!``, styles, column widths and row
heights.  They are **not** presented as real data.  When the real files arrive
they must be parsed by an additional test (see docs/sprint-0-report.md).

They are generated instead of committed as binaries so the structure under test
stays readable and reviewable, and so "the same report one week later" is
expressed as a parameter rather than as a second opaque file.

Datasets required by the specification:

* **A** — weeks ``W31 W32``, months ``Jan…Aug``
* **B** — weeks ``W33 W34``, months ``Jan…Aug``     (window shifted)
* **C** — weeks ``W33 W34``, months ``Jan…Sep``     (a new month appeared)

Run standalone to inspect them: ``python -m tests.fixtures.build_fixtures out/``
"""

from __future__ import annotations

import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
GROUP_FILL = PatternFill("solid", fgColor="DCE7F5")
METRIC_FILL = PatternFill("solid", fgColor="F2F6FC")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
GROUP_FONT = Font(bold=True, color="1E3A5F", size=11, name="Calibri")
TITLE_FONT = Font(bold=True, size=14, color="1E3A5F")
THIN = Side(style="thin", color="B7C4D6")
BORDER = Border(top=THIN, right=THIN, bottom=THIN, left=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

MONTHS = ("Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")


def _style_header(ws, row: int, first_col: int, last_col: int) -> None:
    for col in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _write_row(ws, row: int, first_col: int, values, number_format: str | None = None) -> None:
    for offset, value in enumerate(values):
        cell = ws.cell(row=row, column=first_col + offset, value=value)
        cell.border = BORDER
        cell.alignment = CENTER
        if number_format and isinstance(value, (int, float)):
            cell.number_format = number_format


def _merge_label(ws, row1: int, row2: int, col: int, value: str, *, group: bool) -> None:
    cell = ws.cell(row=row1, column=col, value=value)
    cell.font = GROUP_FONT
    cell.fill = GROUP_FILL if group else METRIC_FILL
    cell.alignment = CENTER
    cell.border = BORDER
    if row2 > row1:
        ws.merge_cells(start_row=row1, start_column=col, end_row=row2, end_column=col)


def _metric_value(rng: random.Random, metric: str) -> object:
    metric = metric.lower()
    if metric.startswith("insp"):
        return rng.randint(3_000, 90_000)
    if metric.startswith("def"):
        return rng.randint(0, 120)
    if metric.startswith("ppm"):
        return round(rng.uniform(50, 900), 1)
    if metric == "target":
        return 300
    return round(rng.uniform(50, 900), 1)


def _number_format(metric: str) -> str:
    metric = metric.lower()
    return "#,##0" if metric.startswith(("insp", "def")) else "#,##0.0"


# --------------------------------------------------------------------------- #
# IQC / OQC weekly matrix — category > subcategory > metric
# --------------------------------------------------------------------------- #
def build_weekly_matrix(
    path: Path,
    *,
    department: str = "IQC",
    years: tuple[str, ...] = ("2025", "2026"),
    months: tuple[str, ...] = MONTHS[:8],
    weeks: tuple[str, ...] = ("W31", "W32"),
    sections: tuple[str, ...] = ("SEC", "TNP", "TECPLAM"),
    subgroups: tuple[str, ...] = ("Total", "TSI", "Packing"),
    metrics: tuple[str, ...] = ("PPM", "Def.", "Insp."),
    seed: int = 7,
) -> Path:
    """The weekly report: three label columns, mixed period columns.

    Nothing about the layout is fixed — the caller decides how many months and
    which weeks exist, which is exactly how the real file changes every week.
    """
    rng = random.Random(seed)
    wb = Workbook()
    ws = wb.active
    ws.title = department

    first_col = 2  # column B — provenance only
    periods = list(years) + list(months) + list(weeks)
    label_cols = 3
    last_col = first_col + label_cols - 1 + len(periods)

    ws.cell(row=1, column=first_col, value=f"{department} — Quality Weekly Report")
    ws.merge_cells(start_row=1, start_column=first_col, end_row=1, end_column=last_col)
    ws.cell(row=1, column=first_col).font = TITLE_FONT
    ws.cell(row=1, column=first_col).alignment = CENTER
    ws.row_dimensions[1].height = 26

    header_row = 2
    _write_row(ws, header_row, first_col, ["Section", "Group", "Metric", *periods])
    _style_header(ws, header_row, first_col, last_col)
    ws.row_dimensions[header_row].height = 20

    row = header_row + 1
    for section in sections:
        section_start = row
        for subgroup in subgroups:
            group_start = row
            for metric in metrics:
                cell = ws.cell(row=row, column=first_col + 2, value=metric)
                cell.border = BORDER
                cell.alignment = LEFT
                if metric.lower().startswith("ppm"):
                    cell.font = Font(bold=True, name="Calibri", size=11)

                values: list[object] = []
                for period in periods:
                    value = _metric_value(rng, metric)
                    if period == years[0] and section == "TECPLAM":
                        value = "NA"  # this section did not exist last year
                    if period == weeks[-1] and metric.lower().startswith("ppm") and subgroup == "Packing":
                        value = "#DIV/0!"  # inspection count was zero
                    values.append(value)
                _write_row(ws, row, first_col + 3, values, number_format=_number_format(metric))
                ws.row_dimensions[row].height = 18
                row += 1
            _merge_label(ws, group_start, row - 1, first_col + 1, subgroup, group=False)
        _merge_label(ws, section_start, row - 1, first_col, section, group=True)

    ws.column_dimensions[get_column_letter(first_col)].width = 14
    ws.column_dimensions[get_column_letter(first_col + 1)].width = 12
    ws.column_dimensions[get_column_letter(first_col + 2)].width = 18
    for col in range(first_col + 3, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10.5

    wb.save(path)
    return path


# --------------------------------------------------------------------------- #
# OQC / TECPLAM — nested merged header (year > month > series)
# --------------------------------------------------------------------------- #
def build_nested_header(path: Path, seed: int = 11) -> Path:
    rng = random.Random(seed)
    wb = Workbook()
    ws = wb.active
    ws.title = "OQC"

    first_col = 2
    months = ["Jun", "Jul", "Aug"]
    series = ["Target", "Result"]
    last_col = first_col + len(months) * len(series)

    ws.cell(row=2, column=first_col, value="Line")
    ws.merge_cells(start_row=2, start_column=first_col, end_row=4, end_column=first_col)
    ws.cell(row=2, column=first_col + 1, value="2026")
    ws.merge_cells(start_row=2, start_column=first_col + 1, end_row=2, end_column=last_col)

    col = first_col + 1
    for month in months:
        ws.cell(row=3, column=col, value=month)
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
        for offset, name in enumerate(series):
            ws.cell(row=4, column=col + offset, value=name)
        col += len(series)

    for row in (2, 3, 4):
        _style_header(ws, row, first_col, last_col)

    for index, line in enumerate(["TECPLAM 1", "TECPLAM 2", "TECPLAM 3"]):
        row = 5 + index
        cell = ws.cell(row=row, column=first_col, value=line)
        cell.border = BORDER
        values: list[object] = [round(rng.uniform(100, 800), 1) for _ in range(len(months) * len(series))]
        values[-1] = "NA"
        _write_row(ws, row, first_col + 1, values, number_format="#,##0.0")

    wb.save(path)
    return path


# --------------------------------------------------------------------------- #
# FIELD — ASR and CASR side by side, each with its own hierarchy
# --------------------------------------------------------------------------- #
def build_field(path: Path, weeks: tuple[str, ...] = ("W31", "W32", "W33"), seed: int = 3) -> Path:
    """ASR and CASR share one raw data file, separated by an empty column."""
    rng = random.Random(seed)
    wb = Workbook()
    ws = wb.active
    ws.title = "FIELD"

    def block(anchor_col: int, indicator: str, number_format: str, percent: bool) -> None:
        label_cols = 3
        last_col = anchor_col + label_cols - 1 + len(weeks)

        ws.cell(row=1, column=anchor_col, value=f"{indicator} — Field Quality")
        ws.merge_cells(start_row=1, start_column=anchor_col, end_row=1, end_column=last_col)
        ws.cell(row=1, column=anchor_col).font = TITLE_FONT
        ws.cell(row=1, column=anchor_col).alignment = CENTER

        _write_row(ws, 2, anchor_col, ["Indicator", "Model", "Metric", *weeks])
        _style_header(ws, 2, anchor_col, last_col)

        row = 3
        indicator_start = row
        for model in ("MX", "Mobile"):
            model_start = row
            for metric in ("Target", "Result"):
                cell = ws.cell(row=row, column=anchor_col + 2, value=metric)
                cell.border = BORDER
                cell.alignment = LEFT
                values: list[object] = []
                for week in weeks:
                    value: object = (
                        round(rng.uniform(0.01, 0.2), 4) if percent else rng.randint(100, 2_000)
                    )
                    if metric == "Target":
                        value = 0.05 if percent else 500
                    if week == weeks[-1] and metric == "Result" and model == "Mobile":
                        value = "#DIV/0!" if indicator == "CASR" else "NA"
                    values.append(value)
                _write_row(ws, row, anchor_col + 3, values, number_format=number_format)
                row += 1
            _merge_label(ws, model_start, row - 1, anchor_col + 1, model, group=False)
        _merge_label(ws, indicator_start, row - 1, anchor_col, indicator, group=True)

        ws.column_dimensions[get_column_letter(anchor_col)].width = 13
        ws.column_dimensions[get_column_letter(anchor_col + 1)].width = 11
        ws.column_dimensions[get_column_letter(anchor_col + 2)].width = 12

    block(2, "ASR", "0.00%", percent=True)
    block(2 + 3 + len(weeks) + 1, "CASR", "#,##0", percent=False)
    wb.save(path)
    return path


# --------------------------------------------------------------------------- #
# FIELD — transposed: periods run down the first column
# --------------------------------------------------------------------------- #
def build_transposed(path: Path, seed: int = 5) -> Path:
    rng = random.Random(seed)
    wb = Workbook()
    ws = wb.active
    ws.title = "FIELD_TRANSPOSED"

    _write_row(ws, 2, 2, ["Period", "Sales", "Defects", "ASR", "CASR"])
    _style_header(ws, 2, 2, 6)
    for index, period in enumerate(["W30", "W31", "W32", "W33"]):
        row = 3 + index
        ws.cell(row=row, column=2, value=period).border = BORDER
        _write_row(
            ws,
            row,
            3,
            [
                rng.randint(10_000, 50_000),
                rng.randint(1, 40),
                round(rng.uniform(0.01, 0.3), 4),
                rng.randint(50, 900),
            ],
        )
    wb.save(path)
    return path


# --------------------------------------------------------------------------- #
# Long / tidy format (the shape of the legacy IQC_2026.xlsx)
# --------------------------------------------------------------------------- #
def build_flat_long(path: Path, seed: int = 13) -> Path:
    rng = random.Random(seed)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    _write_row(ws, 1, 1, ["Ano", "Mes", "Fornecedor", "Defeito", "Produto", "Quantidade", "PPM", "Custo"])
    _style_header(ws, 1, 1, 8)
    row = 2
    for month in range(1, 5):
        for supplier in ("Hyundai", "Bosch", "Denso", "Valeo"):
            for defect in ("Partícula", "Risco"):
                _write_row(
                    ws,
                    row,
                    1,
                    [
                        2026,
                        month,
                        supplier,
                        defect,
                        f"Módulo {rng.choice('ABC')}",
                        rng.randint(60, 400),
                        round(rng.uniform(100, 900), 1),
                        round(rng.uniform(100, 3_000), 2),
                    ],
                )
                row += 1
    wb.save(path)
    return path


# --------------------------------------------------------------------------- #
FIXTURES = {
    # dataset A — the current week
    "iqc_dataset_a.xlsx": lambda p: build_weekly_matrix(p, department="IQC", weeks=("W31", "W32")),
    # dataset B — the window moved, nothing else changed
    "iqc_dataset_b.xlsx": lambda p: build_weekly_matrix(
        p, department="IQC", weeks=("W33", "W34"), seed=8
    ),
    # dataset C — a new month appeared and a metric was added
    "iqc_dataset_c.xlsx": lambda p: build_weekly_matrix(
        p,
        department="IQC",
        months=MONTHS[:9],
        weeks=("W33", "W34"),
        metrics=("PPM", "Def.", "Insp.", "Target"),
        seed=9,
    ),
    "oqc_tecplam.xlsx": build_nested_header,
    "field_asr_casr.xlsx": build_field,
    "field_transposed.xlsx": build_transposed,
    "flat_long.xlsx": build_flat_long,
}


def build_all(out_dir: Path) -> dict[str, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    return {name: builder(out_dir / name) for name, builder in FIXTURES.items()}


if __name__ == "__main__":  # pragma: no cover
    import sys

    target = Path(sys.argv[1] if len(sys.argv) > 1 else "generated")
    for name, path in build_all(target).items():
        print(f"{name}: {path}")
