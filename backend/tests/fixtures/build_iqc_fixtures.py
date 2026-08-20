"""Synthetic IQC workbooks derived from the real one.

The real file (``tests/fixtures/real/RawdataIQC.xlsx``) is a photograph of one
week.  These fixtures reproduce **its structure** — three tables side by side,
merged category cells, unlabelled headline rows, ``SKD``/``CKD`` sub-groups —
with the period axis as a parameter, so "the same report a month later" is an
argument instead of a second opaque file:

======  ===========================================================
A       ``'25 | '26 | 1Q | 2Q | 3Q | Aug``          (today)
B       ``… | Aug | Sep``                            (a month later)
C       ``… | 3Q | 4Q | Aug | Sep | Oct``          (a quarter opens beside its months)
D       ``'25 | '26 | 1Q | 2Q | 3Q | 4Q | Nov | Dec``(the quarter closed)
E       ``… | Nov | Dec | W48``                      (weeks appear)
======  ===========================================================

The same parser code must read all of them.  Values are generated so that
``Rej. Lot / Insp. Lot × 1_000_000`` equals the headline row, which is what the
headline verification checks.
"""

from __future__ import annotations

import random
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIN = Side(style="thin", color="000000")
BORDER = Border(top=THIN, right=THIN, bottom=THIN, left=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
HEADER_FONT = Font(bold=True, size=11)
NAME_FILL = PatternFill("solid", fgColor="D9E1F2")

#: the block layout of the real file: (category, subgroup)
BLOCKS: tuple[tuple[str | None, str | None], ...] = (
    (None, None),  # the unnamed leading block — the table's own total
    ("Imported", None),
    ("Imported", "SKD"),
    ("Imported", "CKD"),
    ("Local", None),
)

#: the metrics written under each headline row
CYCLE = ("Rej. Lot", "Insp. Lot")

PERIODS_A = ("'25", "'26", "1Q", "2Q", "3Q", "Aug")
PERIODS_B = ("'25", "'26", "1Q", "2Q", "3Q", "Aug", "Sep")
#: dataset C keeps the running months *and* shows 4Q, so a quarter and the
#: months inside it live side by side (Sprint 3 §12)
PERIODS_C = ("'25", "'26", "1Q", "2Q", "3Q", "4Q", "Aug", "Sep", "Oct")
PERIODS_D = ("'25", "'26", "1Q", "2Q", "3Q", "4Q", "Nov", "Dec")
PERIODS_E = ("'25", "'26", "1Q", "2Q", "3Q", "4Q", "Nov", "Dec", "W48")


@dataclass
class BlockValues:
    """One block: rejected lots, inspected lots and the PPM they imply."""

    rejected: int
    inspected: int

    @property
    def ppm(self) -> int:
        return round(self.rejected / self.inspected * 1_000_000) if self.inspected else 0


def _block_values(rng: random.Random) -> BlockValues:
    inspected = rng.randint(300, 25_000)
    rejected = rng.randint(0, max(1, inspected // 150))
    return BlockValues(rejected=rejected, inspected=inspected)


def build_iqc(
    path: Path,
    *,
    periods: tuple[str, ...] = PERIODS_A,
    tables: tuple[str, ...] = ("TTL", "SEC", "TNP"),
    blocks: tuple[tuple[str | None, str | None], ...] = BLOCKS,
    seed: int = 42,
) -> Path:
    """Write one IQC workbook with the real structure and the given periods."""
    rng = random.Random(seed)
    wb = Workbook()
    ws = wb.active
    ws.title = "IQC"

    label_cols = 2
    width = label_cols + len(periods)
    first_col = 2  # column B, like the real file
    header_row = 2

    for table_index, name in enumerate(tables):
        anchor = first_col + table_index * (width + 1)  # one empty separator column
        last_col = anchor + width - 1

        # header: the table names itself across the label columns, then periods
        cell = ws.cell(row=header_row, column=anchor, value=name)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.fill = NAME_FILL
        cell.border = BORDER
        ws.merge_cells(
            start_row=header_row,
            start_column=anchor,
            end_row=header_row,
            end_column=anchor + label_cols - 1,
        )
        for offset, period in enumerate(periods):
            period_cell = ws.cell(row=header_row, column=anchor + label_cols + offset, value=period)
            period_cell.font = HEADER_FONT
            period_cell.alignment = CENTER
            period_cell.border = BORDER

        row = header_row + 1
        for category, subgroup in blocks:
            block_start = row
            for line, metric in enumerate((None, *CYCLE)):
                # the headline row carries the sub-group name, or nothing at all
                label = subgroup if (line == 0 and subgroup) else metric
                if label:
                    label_cell = ws.cell(row=row, column=anchor + 1, value=label)
                    label_cell.border = BORDER
                for offset, _period in enumerate(periods):
                    values = _block_values(rng)
                    value = (
                        values.ppm
                        if line == 0
                        else values.rejected
                        if metric == "Rej. Lot"
                        else values.inspected
                    )
                    # keep the three rows of a period consistent with each other
                    if line == 0:
                        _cache[(row, offset)] = values
                    else:
                        cached = _cache.get((block_start, offset))
                        if cached:
                            value = cached.rejected if metric == "Rej. Lot" else cached.inspected
                    data_cell = ws.cell(
                        row=row, column=anchor + label_cols + offset, value=value
                    )
                    data_cell.border = BORDER
                    data_cell.number_format = "#,##0"
                row += 1
            if category:
                # the category is written once and merged over its whole block,
                # exactly as in the real file
                category_cell = ws.cell(row=block_start, column=anchor, value=category)
                category_cell.alignment = CENTER
                category_cell.border = BORDER
                if row - 1 > block_start:
                    ws.merge_cells(
                        start_row=block_start,
                        start_column=anchor,
                        end_row=row - 1,
                        end_column=anchor,
                    )

        ws.column_dimensions[get_column_letter(anchor)].width = 12
        ws.column_dimensions[get_column_letter(anchor + 1)].width = 11
        for col in range(anchor + label_cols, last_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 9

    wb.save(path)
    _cache.clear()
    return path


#: values of a block's headline row, reused by its metric rows
_cache: dict[tuple[int, int], BlockValues] = {}


FIXTURES = {
    "iqc_evolution_a.xlsx": lambda p: build_iqc(p, periods=PERIODS_A, seed=1),
    "iqc_evolution_b.xlsx": lambda p: build_iqc(p, periods=PERIODS_B, seed=2),
    "iqc_evolution_c.xlsx": lambda p: build_iqc(p, periods=PERIODS_C, seed=3),
    "iqc_evolution_d.xlsx": lambda p: build_iqc(p, periods=PERIODS_D, seed=4),
    "iqc_evolution_e.xlsx": lambda p: build_iqc(p, periods=PERIODS_E, seed=5),
}


def build_all(out_dir: Path) -> dict[str, Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    return {name: builder(out_dir / name) for name, builder in FIXTURES.items()}


if __name__ == "__main__":  # pragma: no cover
    import sys

    target = Path(sys.argv[1] if len(sys.argv) > 1 else "generated")
    for name, path in build_all(target).items():
        print(f"{name}: {path}")
