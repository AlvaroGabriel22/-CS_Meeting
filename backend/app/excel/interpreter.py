"""**Layer 2 — Interpreter: raw structure -> semantic structure.**

Decides what the region *means*: where the title is, which rows are headers,
which columns name the rows, what each header token is (year / month / week /
series), and how the row labels nest (category > subcategory > metric).

It produces decisions only — no cells are built here (that is the normalizer's
job, layer 3).  Everything is inferred; a :class:`DepartmentSchema` may be
supplied to raise confidence, but is never required.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field

from openpyxl.utils import get_column_letter

from app.domain.departments import DepartmentSchema, canonical, is_metric_token

from . import hierarchy as H
from . import period_engine as PE
from . import periods as P
from . import values as V
from .model import (
    CellRole,
    ColumnDescriptor,
    PeriodAxis,
    PeriodKind,
    RowDescriptor,
    SemanticType,
    TableShape,
    ValueType,
)
from .raw_model import RawSheet
from .regions import Rect

logger = logging.getLogger(__name__)

MAX_HEADER_ROWS = 6
MAX_LABEL_COLS = 4
#: header rows sparser than this get their group labels forward-filled
SPARSE_HEADER_RATIO = 0.6

#: column names that carry the time dimension in flat/long tables
FLAT_PERIOD_COLUMNS = {
    "year": {"year", "ano", "yr", "연도"},
    "month": {"month", "mes", "월"},
    "week": {"week", "semana", "sem", "주"},
    "day": {"day", "dia", "date", "data", "일"},
}


@dataclass
class TableInterpretation:
    """The semantic reading of one region — the contract between layers 2 and 3."""

    rect: Rect
    title: str | None = None
    title_rows: int = 0
    header_rows: int = 0
    label_cols: int = 0
    columns: list[ColumnDescriptor] = field(default_factory=list)
    rows: list[RowDescriptor] = field(default_factory=list)
    period_axis: PeriodAxis = PeriodAxis.NONE
    shape: TableShape = TableShape.MATRIX
    hierarchy: tuple[str, ...] = ()
    #: label column index -> hierarchy level name
    label_roles: dict[int, str] = field(default_factory=dict)
    #: (row index, column index) -> semantic, for label cells that do not follow
    #: their column's role (a sub-group sitting in the metric column)
    cell_semantics: dict[tuple[int, int], SemanticType] = field(default_factory=dict)
    reporting_year: int | None = None
    warnings: list[str] = field(default_factory=list)
    meta: dict = field(default_factory=dict)

    @property
    def data_start_row(self) -> int:
        """First Excel row that holds measurements."""
        return self.rect.r1 + self.title_rows + self.header_rows


# --------------------------------------------------------------------------- #
# Row / column statistics
# --------------------------------------------------------------------------- #
def _cell_stats(sheet: RawSheet, row: int, c1: int, c2: int) -> tuple[int, int, int]:
    """(non-empty, measurement-like, textual) counts for one row of a region.

    A numeric cell that reads as a period (``2026``, ``8``, ``W32``) is *not*
    counted as a measurement — otherwise a merged year header would look like a
    data row and the whole header band would be missed.
    """
    entries: list[tuple[str, ValueType]] = []
    for col in range(c1, c2 + 1):
        value_type, _, _, _ = V.coerce(sheet.value(row, col))
        if value_type is ValueType.EMPTY:
            continue
        entries.append((sheet.text(row, col), value_type))

    non_empty = len(entries)
    period_hits = sum(1 for text, _ in entries if P.match_token(text) is not None)
    period_row = period_hits >= max(2, int(0.6 * non_empty))

    measurements = textual = 0
    for text, value_type in entries:
        if period_row and P.match_token(text) is not None:
            continue
        if value_type in (ValueType.NUMBER, ValueType.ERROR, ValueType.NA):
            measurements += 1
        elif value_type in (ValueType.TEXT, ValueType.DATE):
            textual += 1
    return non_empty, measurements, textual


def _detect_title(sheet: RawSheet, rect: Rect) -> tuple[str | None, int]:
    """A wide, lonely, non-numeric row at the top of the region is a title."""
    title_parts: list[str] = []
    row = rect.r1
    while row <= rect.r2 and len(title_parts) < 2:
        distinct = {
            sheet.text(row, col) for col in range(rect.c1, rect.c2 + 1) if sheet.is_occupied(row, col)
        }
        non_empty, numeric, _ = _cell_stats(sheet, row, rect.c1, rect.c2)
        if numeric == 0 and len(distinct) == 1 and non_empty >= 1 and rect.rows > 2:
            title_parts.append(distinct.pop())
            row += 1
            continue
        break
    return (" — ".join(title_parts) if title_parts else None), len(title_parts)


def _detect_header_rows(sheet: RawSheet, rect: Rect, first_row: int) -> int:
    """Number of leading rows that describe the data instead of being data."""
    for offset, row in enumerate(range(first_row, rect.r2 + 1)):
        non_empty, numeric, _ = _cell_stats(sheet, row, rect.c1, rect.c2)
        if non_empty and numeric >= max(2, int(0.4 * non_empty)):
            return min(offset, MAX_HEADER_ROWS)
    return 1 if rect.r2 > first_row else 0


def _detect_label_cols(sheet: RawSheet, rect: Rect, data_start: int) -> int:
    """Leading columns that name the rows instead of holding measurements."""
    if data_start > rect.r2:
        return 0
    count = 0
    for col in range(rect.c1, min(rect.c1 + MAX_LABEL_COLS, rect.c2 + 1)):
        non_empty = numeric = 0
        for row in range(data_start, rect.r2 + 1):
            value_type, _, _, _ = V.coerce(sheet.value(row, col))
            if value_type is ValueType.EMPTY:
                continue
            non_empty += 1
            if value_type is ValueType.NUMBER:
                numeric += 1
        if non_empty and numeric >= max(1, int(0.5 * non_empty)):
            break
        count += 1
    return min(count, rect.cols - 1) if rect.cols > 1 else 0


def _header_row_values(sheet: RawSheet, row: int, c1: int, c2: int) -> list[str]:
    """Values of a header row, forward-filled when it holds sparse group labels.

    Excel authors often write ``2026`` once above a block of months instead of
    merging the cells.  Merged ranges are already resolved by the parser, so
    this only covers the un-merged case.
    """
    raw = [sheet.text(row, col) for col in range(c1, c2 + 1)]
    filled_ratio = sum(1 for value in raw if value) / max(len(raw), 1)
    if filled_ratio >= SPARSE_HEADER_RATIO:
        return raw
    out: list[str] = []
    current = ""
    for value in raw:
        if value:
            current = value
        out.append(current)
    return out


# --------------------------------------------------------------------------- #
# Hierarchy: category > subcategory > metric
# --------------------------------------------------------------------------- #
def _label_matrix(
    sheet: RawSheet, rect: Rect, data_start: int, label_cols: int
) -> list[list[H.LabelCell]]:
    """Label cells per data row.

    ``value`` is the label in force (a merged or blank cell inherits the one
    above); ``written`` is True only where the analyst actually typed it, which
    is what tells a new block from a continuation.
    """
    carry = [""] * label_cols
    matrix: list[list[H.LabelCell]] = []
    for row in range(data_start, rect.r2 + 1):
        parts: list[H.LabelCell] = []
        for i in range(label_cols):
            cell = sheet.get(row, rect.c1 + i)
            text = sheet.text(row, rect.c1 + i)
            written = bool(text) and (cell is None or not cell.merged_range or cell.is_merge_anchor)
            if text:
                carry[i] = text
                carry[i + 1 :] = [""] * (label_cols - i - 1)
            parts.append(H.LabelCell(value=text or carry[i], written=written))
        matrix.append(parts)
    return matrix


# --------------------------------------------------------------------------- #
# Main entry point
# --------------------------------------------------------------------------- #
def _corner_table_name(
    sheet: RawSheet, rect: Rect, first_row: int, label_cols: int
) -> str | None:
    """The table's own name, when the header's corner cell carries it.

    In the real IQC sheet the header row starts with ``TTL`` merged across both
    label columns; that is the table talking about itself, not a column title.
    """
    if label_cols < 1:
        return None
    cell = sheet.get(first_row, rect.c1)
    text = sheet.text(first_row, rect.c1)
    if not text or cell is None or not cell.merged_range:
        return None
    if P.match_token(text) or P.match_series(text):
        return None  # a merged period header, not a name
    from openpyxl.utils import range_boundaries

    c1, _r1, c2, _r2 = range_boundaries(cell.merged_range)
    return text if (c2 - c1 + 1) >= label_cols else None



def _is_series_row(values: list[str], row_kind: P.PeriodKind | None) -> bool:
    """True when a header row really splits the columns into series.

    ``Target | Result | Target | Result`` under the months is a series axis.
    ``Simulation | Result | Result | Partial`` is not: two of those words name
    nothing the system knows, which is what a *qualifier* looks like — a note
    about how firm each period's figure is.  The rule is all-or-nothing on
    purpose: a real axis labels its columns consistently, and anything else is
    read as an annotation and left in the table where the analyst wrote it
    (ADR-0045).
    """
    tokens = [
        text
        for text in (str(value).strip() for value in values)
        if text and P.match_token_in_row(text, row_kind) is None
    ]
    if not tokens:
        return False
    return all(P.match_series(token) for token in tokens)


def interpret_region(
    sheet: RawSheet, rect: Rect, schema: DepartmentSchema | None = None
) -> TableInterpretation:
    """Read one region semantically."""
    title, title_rows = _detect_title(sheet, rect)
    first_row = rect.r1 + title_rows
    header_rows = _detect_header_rows(sheet, rect, first_row)
    data_start = first_row + header_rows
    label_cols = _detect_label_cols(sheet, rect, data_start)
    corner_name = _corner_table_name(sheet, rect, first_row, label_cols) if header_rows else None
    if corner_name and not title:
        title = corner_name

    interpretation = TableInterpretation(
        rect=rect,
        title=title,
        title_rows=title_rows,
        header_rows=header_rows,
        label_cols=label_cols,
    )

    header_values = [
        _header_row_values(sheet, first_row + i, rect.c1, rect.c2) for i in range(header_rows)
    ]
    row_kinds = [P.row_period_kind(row[label_cols:]) for row in header_values]
    series_rows = [
        _is_series_row(row[label_cols:], kind) for row, kind in zip(header_values, row_kinds)
    ]

    # ---------------- columns --------------------------------------------- #
    for offset, col in enumerate(range(rect.c1, rect.c2 + 1)):
        header_path = tuple(row[offset] for row in header_values if row[offset])
        is_label = offset < label_cols
        period, series_type = (None, None)
        if not is_label and header_rows:
            period, series_type = P.build_period(
                [row[offset] for row in header_values], row_kinds, series_rows
            )
        if is_label:
            semantic = SemanticType.LABEL
        elif period:
            semantic = SemanticType.PERIOD
        elif series_type:
            semantic = SemanticType.SERIES
        else:
            semantic = SemanticType.UNKNOWN
        interpretation.columns.append(
            ColumnDescriptor(
                index=offset,
                source_column=get_column_letter(col),
                header_path=header_path,
                # a period column is named by its period: the header may also
                # carry a qualifier (``Simulation``), and that word names the
                # figure's firmness, not the column
                label=(
                    period.label
                    if period and period.label
                    else (header_path[-1] if header_path else get_column_letter(col))
                ),
                period=period,
                series_type=series_type,
                semantic=semantic,
                is_label_column=is_label,
                width=sheet.col_widths.get(col),
            )
        )

    # ---------------- rows & hierarchy ------------------------------------ #
    matrix = _label_matrix(sheet, rect, data_start, label_cols)
    labels = H.analyze(matrix, label_cols, schema)
    interpretation.label_roles = labels.roles
    interpretation.hierarchy = labels.hierarchy
    interpretation.warnings.extend(labels.warnings)
    interpretation.meta.update(labels.meta)

    for offset, row in enumerate(range(rect.r1, rect.r2 + 1)):
        is_header = offset < title_rows + header_rows
        descriptor = RowDescriptor(
            index=offset,
            source_row=row,
            is_header_row=is_header,
            height=sheet.row_heights.get(row),
            semantic=SemanticType.TITLE
            if offset < title_rows
            else SemanticType.PERIOD
            if is_header
            else SemanticType.VALUE,
        )
        if not is_header and label_cols:
            data_index = offset - (title_rows + header_rows)
            cells = matrix[data_index]
            row_labels = labels.rows[data_index]
            descriptor.label_path = tuple(cell.value for cell in cells if cell.value)
            descriptor.label = descriptor.label_path[-1] if descriptor.label_path else ""
            descriptor.level = max(len(descriptor.label_path) - 1, 0)
            descriptor.category = row_labels.category
            descriptor.subcategory = row_labels.subcategory
            descriptor.metric = row_labels.metric
            descriptor.series_type = row_labels.series_type
            descriptor.block = row_labels.block
            descriptor.inferred = row_labels.inferred

            if descriptor.metric:
                descriptor.semantic = SemanticType.METRIC
            elif descriptor.series_type:
                descriptor.semantic = SemanticType.SERIES
            elif descriptor.subcategory:
                descriptor.semantic = SemanticType.SUBCATEGORY
            elif descriptor.category:
                descriptor.semantic = SemanticType.CATEGORY

            # a sub-group label written in the metric column is a subcategory
            metric_col = next(
                (index for index, role in labels.roles.items() if role == "metric"), None
            )
            if metric_col is not None and cells[metric_col].written:
                if cells[metric_col].value in labels.subgroups:
                    interpretation.cell_semantics[(offset, metric_col)] = SemanticType.SUBCATEGORY
        interpretation.rows.append(descriptor)

    # ---------------- period engine --------------------------------------- #
    resolution = PE.resolve([column.period for column in interpretation.columns if column.period])
    if resolution.periods:
        resolved = iter(resolution.periods)
        for column in interpretation.columns:
            if column.period:
                column.period = next(resolved)
        interpretation.reporting_year = resolution.reporting_year
        interpretation.warnings.extend(resolution.warnings)

    # ---------------- period axis ----------------------------------------- #
    column_periods = [column.period for column in interpretation.columns if column.period]
    if len(column_periods) >= 2:
        interpretation.period_axis = PeriodAxis.COLUMNS
    elif label_cols:
        first_labels = [
            row.label_path[0] for row in interpretation.rows if not row.is_header_row and row.label_path
        ]
        if P.looks_like_period_sequence(first_labels):
            interpretation.period_axis = PeriodAxis.ROWS
            kind = P.row_period_kind(first_labels)
            for descriptor in interpretation.rows:
                if descriptor.is_header_row or not descriptor.label_path:
                    continue
                period, _ = P.build_period(descriptor.label_path, [kind] * len(descriptor.label_path))
                descriptor.period = period
                if period:
                    # the row *is* a period; it is not a metric or a category
                    descriptor.semantic = SemanticType.PERIOD
                    descriptor.category = descriptor.subcategory = descriptor.metric = None
            row_resolution = PE.resolve([row.period for row in interpretation.rows if row.period])
            if row_resolution.periods:
                resolved_rows = iter(row_resolution.periods)
                for descriptor in interpretation.rows:
                    if descriptor.period:
                        descriptor.period = next(resolved_rows)
                interpretation.reporting_year = row_resolution.reporting_year
            interpretation.hierarchy = ()
            interpretation.label_roles = {}

    # ---------------- shape ------------------------------------------------ #
    data_rows = rect.r2 - data_start + 1
    if interpretation.period_axis is not PeriodAxis.NONE:
        interpretation.shape = TableShape.MATRIX
    elif header_rows <= 1 and data_rows >= 4 and rect.cols >= 3:
        interpretation.shape = TableShape.FLAT
        interpretation.meta["flatPeriodColumns"] = _flat_period_columns(interpretation)
    elif rect.rows < 2 or rect.cols < 2:
        interpretation.shape = TableShape.FRAGMENT
    else:
        interpretation.shape = TableShape.MATRIX
        interpretation.warnings.append("no_period_detected")

    years = {
        column.period.year
        for column in interpretation.columns
        if column.period and column.period.year
    }
    corner = interpretation.columns[0].header_path[-1] if (
        interpretation.columns and interpretation.columns[0].header_path
    ) else None
    interpretation.meta.update(
        {
            "contextYear": max(years) if years else None,
            "cornerLabel": corner,
            "tableName": corner_name,
            "reportingYear": interpretation.reporting_year,
            "dataStartRow": data_start,
            "titleRows": title_rows,
            "labelRoles": {str(index): role for index, role in labels.roles.items()},
            "schema": schema.code if schema else None,
        }
    )
    logger.debug(
        "interpreted %s: header=%d labels=%d hierarchy=%s axis=%s",
        rect.a1,
        header_rows,
        label_cols,
        interpretation.hierarchy,
        interpretation.period_axis.value,
    )
    return interpretation


def _flat_period_columns(interpretation: TableInterpretation) -> dict[str, int]:
    """Map ``year``/``month``/``week``/``day`` to column indexes of a flat table."""
    found: dict[str, int] = {}
    for column in interpretation.columns:
        name = canonical(column.label)
        for key, aliases in FLAT_PERIOD_COLUMNS.items():
            if name in aliases and key not in found:
                found[key] = column.index
    return found


__all__ = ["TableInterpretation", "interpret_region", "PeriodKind"]
