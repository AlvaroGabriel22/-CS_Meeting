"""**Layer 1 — Parser: Excel file -> raw parsed structure.**

The only module in the system allowed to import openpyxl.  It reads what is
there and nothing else; meaning is added later by the interpreter.

The workbook is opened twice: with cached values (what the analyst sees,
including ``#DIV/0!``) and with formulas (kept for traceability).
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from .model import DEFAULT_STYLE, CellStyle
from .raw_model import RawCell, RawSheet, RawWorkbook

logger = logging.getLogger(__name__)


def parse_workbook(path: str | Path) -> RawWorkbook:
    """Read every visible sheet of ``path`` into a :class:`RawWorkbook`."""
    path = Path(path)
    wb_values = load_workbook(path, data_only=True)
    wb_formulas = load_workbook(path, data_only=False)
    workbook = RawWorkbook(filename=path.name)

    try:
        for worksheet in wb_values.worksheets:
            if worksheet.sheet_state != "visible":
                logger.debug("skipping hidden sheet %r", worksheet.title)
                continue
            workbook.sheets.append(_read_sheet(worksheet, wb_formulas[worksheet.title]))
    finally:
        wb_values.close()
        wb_formulas.close()

    if not workbook.sheets:
        workbook.warnings.append("no_visible_sheet")
    return workbook


def _read_sheet(ws: Any, ws_formulas: Any) -> RawSheet:
    merge_map: dict[tuple[int, int], tuple[tuple[int, int], str]] = {}
    for rng in ws.merged_cells.ranges:
        anchor = (rng.min_row, rng.min_col)
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                merge_map[(row, col)] = (anchor, str(rng))

    sheet = RawSheet(
        name=ws.title,
        max_row=ws.max_row or 1,
        max_col=ws.max_column or 1,
        merged_ranges=[str(rng) for rng in ws.merged_cells.ranges],
    )

    for row in ws.iter_rows():
        for cell in row:
            coord = (cell.row, cell.column)
            merge = merge_map.get(coord)
            anchor_coord = merge[0] if merge else coord
            source = ws.cell(row=anchor_coord[0], column=anchor_coord[1])
            formula_cell = ws_formulas.cell(row=anchor_coord[0], column=anchor_coord[1])
            formula = (
                formula_cell.value
                if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")
                else None
            )
            if source.value is None and formula is None and not merge:
                if _extract_style(cell).is_default:
                    continue  # truly empty: keep the sheet sparse
            sheet.cells[coord] = RawCell(
                row=cell.row,
                col=cell.column,
                value=source.value,
                formula=formula,
                number_format=source.number_format,
                style=_extract_style(source),
                merged_range=merge[1] if merge else None,
                is_merge_anchor=bool(merge) and coord == anchor_coord,
            )

    for letter, dim in ws.column_dimensions.items():
        if not dim.width:
            continue
        first = dim.min or column_index_from_string(letter)
        last = dim.max or first
        for col in range(int(first), int(last) + 1):
            sheet.col_widths[col] = float(dim.width)
    for index, dim in ws.row_dimensions.items():
        if dim.height:
            sheet.row_heights[int(index)] = float(dim.height)

    logger.debug(
        "read sheet %r: %d cells, %d merged range(s)",
        sheet.name,
        len(sheet.cells),
        len(sheet.merged_ranges),
    )
    return sheet


# --------------------------------------------------------------------------- #
# Style extraction (visual metadata — never drives data logic)
# --------------------------------------------------------------------------- #
def _color(obj: Any) -> str | None:
    """Extract an ``RRGGBB`` string from an openpyxl colour, if it has one."""
    rgb = getattr(obj, "rgb", None)
    if isinstance(rgb, str) and len(rgb) in (6, 8):
        value = rgb[-6:].upper()
        return None if value == "000000" and getattr(obj, "type", "") == "theme" else value
    return None


def _extract_style(cell: Any) -> CellStyle:
    font = cell.font
    fill = cell.fill
    align = cell.alignment
    border = cell.border
    borders = tuple(
        side
        for side in ("top", "right", "bottom", "left")
        if getattr(getattr(border, side, None), "style", None)
    )
    fill_color = None
    if fill is not None and getattr(fill, "patternType", None) == "solid":
        fill_color = _color(fill.fgColor)
        if fill_color == "FFFFFF":
            fill_color = None
    style = CellStyle(
        bold=bool(font and font.bold),
        italic=bool(font and font.italic),
        underline=bool(font and font.underline),
        font_size=float(font.size) if font and font.size else None,
        font_name=font.name if font else None,
        font_color=_color(font.color) if font and font.color else None,
        fill_color=fill_color,
        align_h=align.horizontal if align else None,
        align_v=align.vertical if align else None,
        wrap=bool(align and align.wrap_text),
        borders=borders,
    )
    return DEFAULT_STYLE if style.is_default else style
