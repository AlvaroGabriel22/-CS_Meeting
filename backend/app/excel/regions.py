"""Table region detection.

A sheet may hold one table or several side by side.  Instead of being told
"the table lives at B2:Q40", the parser *finds* the blocks: it trims empty
borders and splits recursively on fully empty rows/columns (a guillotine cut).
The resulting range is then recorded as provenance — so ``B2:Q40`` still shows
up in the UI for traceability, but nothing depends on it.
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.utils import get_column_letter

from .raw_model import RawSheet

MAX_REGIONS = 40
MIN_CELLS = 2


@dataclass(frozen=True)
class Rect:
    r1: int
    c1: int
    r2: int
    c2: int

    @property
    def rows(self) -> int:
        return self.r2 - self.r1 + 1

    @property
    def cols(self) -> int:
        return self.c2 - self.c1 + 1

    @property
    def a1(self) -> str:
        return f"{get_column_letter(self.c1)}{self.r1}:{get_column_letter(self.c2)}{self.r2}"

    def __bool__(self) -> bool:
        return self.r1 <= self.r2 and self.c1 <= self.c2


def _row_empty(grid: RawSheet, row: int, c1: int, c2: int) -> bool:
    return not any(grid.is_occupied(row, col) for col in range(c1, c2 + 1))


def _col_empty(grid: RawSheet, col: int, r1: int, r2: int) -> bool:
    return not any(grid.is_occupied(row, col) for row in range(r1, r2 + 1))


def trim(grid: RawSheet, rect: Rect) -> Rect | None:
    """Shrink a rectangle until every border row/column holds something."""
    r1, c1, r2, c2 = rect.r1, rect.c1, rect.r2, rect.c2
    while r1 <= r2 and _row_empty(grid, r1, c1, c2):
        r1 += 1
    while r2 >= r1 and _row_empty(grid, r2, c1, c2):
        r2 -= 1
    while c1 <= c2 and _col_empty(grid, c1, r1, r2):
        c1 += 1
    while c2 >= c1 and _col_empty(grid, c2, r1, r2):
        c2 -= 1
    rect = Rect(r1, c1, r2, c2)
    return rect if rect else None


def _split(grid: RawSheet, rect: Rect, out: list[Rect], depth: int = 0) -> None:
    if depth > 12 or len(out) >= MAX_REGIONS:
        out.append(rect)
        return
    trimmed = trim(grid, rect)
    if trimmed is None:
        return

    for row in range(trimmed.r1 + 1, trimmed.r2):
        if _row_empty(grid, row, trimmed.c1, trimmed.c2):
            _split(grid, Rect(trimmed.r1, trimmed.c1, row - 1, trimmed.c2), out, depth + 1)
            _split(grid, Rect(row + 1, trimmed.c1, trimmed.r2, trimmed.c2), out, depth + 1)
            return

    for col in range(trimmed.c1 + 1, trimmed.c2):
        if _col_empty(grid, col, trimmed.r1, trimmed.r2):
            _split(grid, Rect(trimmed.r1, trimmed.c1, trimmed.r2, col - 1), out, depth + 1)
            _split(grid, Rect(trimmed.r1, col + 1, trimmed.r2, trimmed.c2), out, depth + 1)
            return

    out.append(trimmed)


def find_regions(grid: RawSheet) -> list[Rect]:
    """Return every non-empty block of the sheet, in reading order."""
    bounds = Rect(1, 1, max(grid.max_row, 1), max(grid.max_col, 1))
    regions: list[Rect] = []
    _split(grid, bounds, regions)
    regions = [r for r in regions if r and r.rows * r.cols >= MIN_CELLS]
    return sorted(regions, key=lambda r: (r.r1, r.c1))
